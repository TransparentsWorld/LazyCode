import copy
import logging
from datetime import datetime
import configparser
import codecs
import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
from pandas import DataFrame, ExcelWriter
from xlwings import Book, App, Sheet

from OfficeJob.src.GenerateSupervisionJobDataReport.Util import *
from OfficeJob.src.OfficeUtil import *


class GenerateWeeklyReport:
    '''
    生成督察支队每周数据周报
    '''

    def get_dir_all_file(self, dir_path: str) -> list:
        """
        获取某个文件夹下的所有文件路径，不会获取子文件夹下的文件路径
        :param dir_path: 需要遍历的文件夹的路径
        :return: 返回文件夹的路径下的所有文件的路径
        """
        file_paths: list = os.listdir(dir_path)
        temp_paths: list = []
        for file_path in file_paths:
            temp_paths.append(os.path.join(dir_path, file_path))
        return temp_paths

    def get_sheet_last_row(self, wb: Book, sheet: Sheet) -> int:
        '''
        获取工作表最后一行
        :param wb: 工作簿
        :param sheet: 工作表
        :return: 返回最后一行
        '''
        last_cell = wb.sheets[sheet].used_range.last_cell
        last_row = last_cell.row + 1
        return last_row

    def get_sheet_last_column(self, wb: Book, sheet: Sheet) -> int:
        '''
        获取工作表最后一列
        :param wb: 工作簿
        :param sheet: 工作表
        :return: 返回最后一列
        '''
        last_cell = wb.sheets[sheet].used_range.last_cell
        last_column = last_cell.column
        return last_column

    def merge_workbook(self, weekly_summary_wb: Book, files_dir_path: str) -> None:
        '''
        合并各县区上报的数据
        :param weekly_summary_wb: 每周汇总数据表
        :param files_dir_path: 各县区上报表格所在的文件夹路径
        :return: 无返回值
        '''
        sub_wb: Book = None  # 用于汇总的分表
        data_sheets: list = ["举报投诉核查工作", "维权工作", "督察工作情况"]  # 工作簿中的三张工作表

        excel_paths = self.get_dir_all_file(files_dir_path)
        # 循环访问文件夹中的工作簿
        print("正在合并")
        for excel_path in excel_paths:
            try:
                sub_wb: Book = xw.Book(excel_path)
                print("-->" + excel_path)
                # 循环访问工作簿中的三张工作表
                for data_sheet in data_sheets:
                    last_row: int = self.get_sheet_last_row(weekly_summary_wb, data_sheet)
                    last_column: int = self.get_sheet_last_column(weekly_summary_wb, data_sheet)
                    # 通过复制粘贴来合并工作表
                    weekly_summary_wb.sheets[data_sheet].range((last_row, 1), (last_row, last_column)).value = \
                        sub_wb.sheets[data_sheet].range((3, 1), (3, last_column)).value
                    last_row = last_row + 1
            except BaseException as e:
                logging.exception(e)
            finally:
                weekly_summary_wb.save()
                sub_wb.close()
        print('合并完毕')

    def dataframe_append_excel(self, df: DataFrame, all_summary_path: str, sheetname: str) -> None:
        """
        将pandas中的dataframe数据类型追加到excel工作表中。

        :param all_summary_path: 全部数据汇总工作簿路径
        :param df: 需要追加的dataframe
        :return: None
        """
        book = load_workbook(all_summary_path)
        writer: ExcelWriter = ExcelWriter(all_summary_path, engine='openpyxl', date_format="yyyy-mm-dd")
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}

        df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=True,
                    header=False)
        writer.save()

    def deal_data_complaint(self, weekly_summary_path: str, sheet: str, last_column: int, begin_date: str,
                            end_date: str) -> DataFrame:
        '''
        从“举报投诉核查工作表”中提取数据，并进行修正。
        :param weekly_summary_path: 每周数据汇总表所在路径
        :param sheet: 工作簿中工作表的名字
        :param last_column: 工作表的最后一样
        :return: 返回处理后的DataFrame
        '''
        df: DataFrame = pd.read_excel(io=weekly_summary_path, engine='openpyxl', sheet_name=sheet, index_col=0)
        df = df.fillna(0)  # 将excel表中的NaN值替换成0

        # 数据更正
        df['total_deal_case'] = df.apply(lambda x: x['real'] + x['fake'],
                                         axis=1)  # 总共处理案件数 = 属实案件数 + 不属实案件数
        df['total_deal_person'] = df.apply(
            lambda x: x['notice'] + x['confine'] +  # 总共处理人数 = 通报 + 禁闭 + 刑事处分 + 移送纪委 + 开出协警数
                      x['suspension'] + x['discipline'] + x['fired_auxiliary_cop'], axis=1)
        df.loc['Row_sum'] = df.iloc[:, 2:last_column].apply(
            lambda x: x.sum())  # 去掉前三列数据，剩下的数据求和，将值添加到最后一行
        # 设置日期
        df['begin_date'] = begin_date
        df['end_date'] = end_date

        # 设置起始日期的格式为"yyyy-mm-dd"
        # df['begin_date'] = df['begin_date'].apply(lambda x: x.strftime('%Y-%m-%d'))
        # df['end_date'] = df['end_date'].apply(lambda x: x.strftime('%Y-%m-%d'))
        return df

    def deal_data_rights_protection(self, weekly_summary_path: str, sheet: str, last_column: int, begin_date: str,
                                    end_date: str) -> DataFrame:
        '''
        从“维权工作表”中提取数据，并进行修正。
        :param weekly_summary_path: 每周数据汇总表所在路径
        :param sheet: 工作簿中工作表的名字
        :param last_column: 工作表的最后一样
        :return: 返回处理后的DataFrame
        '''
        df: DataFrame = pd.read_excel(io=weekly_summary_path, engine='openpyxl', sheet_name=sheet, index_col=0)
        df = df.fillna(0)  # 将excel表中的NaN值替换成0

        # 设置起始日期的格式为"yyyy-mm-dd"
        df['begin_date'] = df['begin_date'].apply(lambda x: x.strftime('%Y-%m-%d'))
        df['end_date'] = df['end_date'].apply(lambda x: x.strftime('%Y-%m-%d'))
        # 数据求和
        df['total_accept'] = df.apply(
            lambda x: x['assault_cop'] + x['intimidate'] + x['pull'] + x['false_accusation'] + x['containment'] + x[
                'malicious_complaints'] + x['other_accept'], axis=1)  # 办理维权案件总和
        df['total_deal'] = df.apply(
            lambda x: x['criminal_responsibility'] + x['public_security_punishment'] + x['cure'] + x['protect'] + x[
                'on_site_disposal'] + x['condolence'] + x['other_deal'] + x['solatium'], axis=1)  # 维权案件处置情况总和
        df['total_victim'] = df.apply(lambda x: x['victim_cop'] + x['victim_auxiliary_cop'], axis=1)  # 受侵害人员总和

        df.loc['Row_sum'] = df.iloc[:, 2:last_column].apply(
            lambda x: x.sum())  # 去掉前三列数据，剩下的数据求和，将值添加到最后一行

        # 设置日期
        df['begin_date'] = begin_date
        df['end_date'] = end_date

        return df

    def deal_data_supervision(self, weekly_summary_path: str, sheet: str, last_column: int, begin_date: str,
                              end_date: str) -> DataFrame:
        '''
        从“督察情况工作表”中提取数据，并进行修正。
        :param weekly_summary_path: 每周数据汇总表所在路径
        :param sheet: 工作簿中工作表的名字
        :param last_column: 工作表的最后一样
        :return: 返回处理后的DataFrame
        '''
        df: DataFrame = pd.read_excel(io=weekly_summary_path, engine='openpyxl', sheet_name=sheet, index_col=0)
        df = df.fillna(0)  # 将excel表中的NaN值替换成0

        # 设置起始日期的格式为"yyyy-mm-dd"
        df['begin_date'] = df['begin_date'].apply(lambda x: x.strftime('%Y-%m-%d'))
        df['end_date'] = df['end_date'].apply(lambda x: x.strftime('%Y-%m-%d'))

        # 数据修改
        df['net_inspect_point'] = df.apply(lambda x: max(x['net_inspect_unit'], x['net_inspect_point']), axis=1)

        df.loc['Row_sum'] = df.iloc[:, 2:last_column].apply(
            lambda x: x.sum())  # 去掉前三列数据，剩下的数据求和，将值添加到最后一行

        # 设置日期
        df['begin_date'] = begin_date
        df['end_date'] = end_date
        return df

    def generate_render_dict(self, df: DataFrame, last_row: int, last_column: int, index_name: str) -> dict:
        '''
        生成每周全市和市局汇总数据的字典，用于下一步模板渲染。
        :param df: 包含数据的DataFrame
        :param last_row: 工作表的最后一行
        :param last_column: 工作表的最后一列
        :param index_name: 毕节市公安局数据所在行的行名
        :return: 返回全市数据和市局市局数据的字典
        '''
        all_sum = df.iloc[last_row - 2:last_row - 1, 2:last_column].astype('int').to_dict(orient="records")[
            0]  # 将dataframe最后的汇总数据转换成字典返回
        df.drop(index='Row_sum', inplace=True)  # 删除汇总行
        # 生成每周毕节数据的字典，用于下一步模板渲染
        series_report_bj = df.loc[index_name]
        bj_sum: dict = series_report_bj[2:last_column].astype('int').to_dict()
        bj_sum = Util.modify_dict_key(bj_sum)
        all_sum.update(bj_sum)
        return all_sum

    def time_format_convert(date: str) -> str:
        '''
        将英文格式的日期转换成中文，如2021-5-5转2021年5月5日
        :return: 返回转换后的日期
        '''
        date = datetime.strptime(date, "%Y-%m-%d")  # 字符串转datatime
        date = date.strftime('%Y{y}%m{m}%d{d}').format(y='年', m='月', d='日')  # 格式化为中文日期
        return date


if __name__ == '__main__':
    print("全部数据汇总表只能是xlsx格式")
    excels_dir_path: str = input('请输入各县区表格所在目录路径：')
    begin_date: str = input('请输入起始时间(如2021-4-13):')
    end_date: str = input('请输入结束时间(如2021-5-13):')

    report_filename: str = Util.generate_report_filename(begin_date, end_date)
    date: str = Util.generate_date_in_weekly_report(begin_date, end_date)

    # 从配置文件读取文件路径
    cf = configparser.ConfigParser()
    with codecs.open('./config.ini', 'r', encoding='utf-8') as f:
        cf.read_file(f)
    secs = cf.sections()
    all_summary_path: str = cf.get('Util', 'all_summary_path')  # 所有汇总数据存放的文件夹
    word_template_path: str = cf.get('Util', 'word_template_path')  # 用于生成报告的模板文件
    weekly_report: str = cf.get('Util', 'report')
    excel_weekly_summary_template_path: str = cf.get('GenerateWeeklyReport', 'excel_weekly_summary_template_path')
    excel_weekly_summary_dir: str = cf.get('GenerateWeeklyReport', 'excel_weekly_summary_dir')

    word_target_path: str = os.path.join(weekly_report, '督察主要业务数据' + date) + '.docx'
    excel_weekly_summary_path: str = os.path.join(excel_weekly_summary_dir, report_filename) + '.xlsx'  # 每周汇总数据存放的文件夹

    sheets: list = ["举报投诉核查工作", "维权工作", "督察工作情况"]
    app: App = None
    weekly_summary_wb: Book = None

    gw: GenerateWeeklyReport = GenerateWeeklyReport()
    copyfile(excel_weekly_summary_template_path, excel_weekly_summary_path)

    try:
        app = xw.App(visible=False)
        app.display_alerts = False
        app.screen_updating = False

        weekly_summary_wb: Book = xw.Book(excel_weekly_summary_path)
        gw.merge_workbook(weekly_summary_wb, excels_dir_path)

        # 举报投诉工作表
        last_column = gw.get_sheet_last_column(weekly_summary_wb, sheets[0])
        last_row = gw.get_sheet_last_row(weekly_summary_wb, sheets[0])
        df_complaint = gw.deal_data_complaint(excel_weekly_summary_path, sheets[0], last_column, begin_date, end_date)
        complaint_render: dict = gw.generate_render_dict(df_complaint, last_row, last_column, '市局')
        gw.dataframe_append_excel(df_complaint, all_summary_path, sheets[0])

        # 维权工作表
        last_column = gw.get_sheet_last_column(weekly_summary_wb, sheets[1])
        last_row = gw.get_sheet_last_row(weekly_summary_wb, sheets[1])
        df_rights_protection = gw.deal_data_rights_protection(excel_weekly_summary_path, sheets[1], last_column,
                                                              begin_date, end_date)
        rights_protection_render: dict = gw.generate_render_dict(df_rights_protection, last_row, last_column, '市局')
        gw.dataframe_append_excel(df_rights_protection, all_summary_path, sheets[1])

        # 督察工作表
        last_column = gw.get_sheet_last_column(weekly_summary_wb, sheets[2])
        last_row = gw.get_sheet_last_row(weekly_summary_wb, sheets[2])
        df_supervision = gw.deal_data_supervision(excel_weekly_summary_path, sheets[2], last_column, begin_date,
                                                  end_date)
        supervision_render: dict = gw.generate_render_dict(df_supervision, last_row, last_column, '市局')
        gw.dataframe_append_excel(df_supervision, all_summary_path, sheets[2])

        # 合并所有生成报告的数据到同一个字典中
        report_data: dict = copy.deepcopy(complaint_render)
        report_data.update(rights_protection_render)
        report_data.update(supervision_render)
        report_data['date'] = date

        # 生成word报告
        Util.render_docx(report_data, word_template_path, word_target_path)
        print('已生成报告')
    except BaseException as e:
        logging.exception(e)
    finally:
        weekly_summary_wb.save()
        weekly_summary_wb.close()
        app.quit()
